import os, json, io
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, abort, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from dotenv import load_dotenv

load_dotenv() 

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'fallback-dev-key-only')

# --- DATABASE CONFIGURATION ---
env = os.environ.get('FLASK_ENV', 'development').lower()
custom_db_url = os.environ.get('DATABASE_URL')

if custom_db_url:
    app.config['SQLALCHEMY_DATABASE_URI'] = custom_db_url
elif env == 'production':
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///production_tiffin_cpac.db'
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///dev_tiffin_cpac.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'home'

# --- MODELS ---
class Teacher(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=True) 
    title = db.Column(db.String(10), nullable=False)
    first_name = db.Column(db.String(100), nullable=False)
    last_name = db.Column(db.String(100), nullable=False)

    @property
    def role(self): return 'teacher'
    def get_id(self): return f"teacher_{self.id}"

class Cohort(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    start_year = db.Column(db.Integer, nullable=False)
    end_year = db.Column(db.Integer, nullable=False)
    students = db.relationship('Student', backref='cohort', lazy=True, cascade="all, delete-orphan")

class Student(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(100), nullable=False)
    last_name = db.Column(db.String(100), nullable=False)
    
    # New Class Tracking Fields
    y12_class = db.Column(db.String(20), nullable=True)
    y13_class = db.Column(db.String(20), nullable=True)
    
    cohort_id = db.Column(db.Integer, db.ForeignKey('cohort.id'), nullable=False)
    assessments = db.relationship('Assessment', backref='student_ref', lazy=True, cascade="all, delete-orphan")

    @property
    def role(self): return 'student'
    def get_id(self): return f"student_{self.id}"

class Skill(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(10), unique=True, nullable=False)
    description = db.Column(db.String(500), nullable=False)

class Practical(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    description = db.Column(db.String(500), nullable=True)

practical_skills = db.Table('practical_skills',
    db.Column('practical_id', db.Integer, db.ForeignKey('practical.id', ondelete="CASCADE"), primary_key=True),
    db.Column('skill_id', db.Integer, db.ForeignKey('skill.id', ondelete="CASCADE"), primary_key=True)
)
Practical.skills = db.relationship('Skill', secondary=practical_skills, lazy='subquery', backref=db.backref('practicals', lazy=True))

class Assessment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id', ondelete="CASCADE"), nullable=False)
    skill_id = db.Column(db.Integer, db.ForeignKey('skill.id', ondelete="CASCADE"), nullable=False)
    practical_id = db.Column(db.Integer, db.ForeignKey('practical.id', ondelete="CASCADE"), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('teacher.id'), nullable=False)
    date_signed = db.Column(db.DateTime, default=datetime.utcnow)

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id', ondelete="CASCADE"), nullable=False)
    practical_id = db.Column(db.Integer, db.ForeignKey('practical.id', ondelete="CASCADE"), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('teacher.id'), nullable=False)
    is_present = db.Column(db.Boolean, default=False, nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

# --- AUTH & ACCESS CONTROL ---
@login_manager.user_loader
def load_user(user_id):
    if user_id.startswith('teacher_'): return Teacher.query.get(int(user_id.split('_')[1]))
    elif user_id.startswith('student_'): return Student.query.get(int(user_id.split('_')[1]))
    return None

def teacher_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'teacher': abort(403)
        return f(*args, **kwargs)
    return decorated_function

# --- TIME LOGIC & UTILS ---
def get_academic_end_year():
    now = datetime.utcnow()
    return now.year if now.month < 9 else now.year + 1

def get_cohort_title(end_year):
    acad_end = get_academic_end_year()
    if end_year == acad_end: return "Year 13"
    elif end_year == acad_end + 1: return "Year 12"
    elif end_year < acad_end:
        diff = acad_end - end_year
        return "Left last year" if diff == 1 else f"Left {diff} years ago"
    return f"Future (Starts {end_year-2})"

def get_cohort_status(cohort):
    """Returns True if the cohort has entered Year 13 (After Aug 1st of their second year)"""
    now = datetime.utcnow()
    transition_date = datetime(cohort.start_year + 1, 8, 1)
    return now >= transition_date

def parse_student_line(line, format_type):
    """Helper to parse bulk student additions"""
    fname, lname, cname = "", "", ""
    if format_type == "First, Last, Class":
        parts = [p.strip() for p in line.split(',')]
        if len(parts) >= 3: fname, lname, cname = parts[0], parts[1], parts[2]
        elif len(parts) == 2: fname, lname = parts[0], parts[1]
    elif format_type == "Last, First, Class":
        parts = [p.strip() for p in line.split(',')]
        if len(parts) >= 3: lname, fname, cname = parts[0], parts[1], parts[2]
        elif len(parts) == 2: lname, fname = parts[0], parts[1]
    elif format_type == "Last First Class":
        parts = line.split()
        if len(parts) >= 3:
            cname = parts[-1]
            lname = parts[0]
            fname = " ".join(parts[1:-1])
        elif len(parts) == 2:
            lname = parts[0]
            fname = parts[1]
    return fname, lname, cname

# --- ROUTES ---
@app.route('/')
def home():
    if current_user.is_authenticated:
        if current_user.role == 'teacher': return redirect(url_for('dashboard'))
        else: return redirect(url_for('student_view', id=current_user.id))
    return render_template('home.html')

@app.route('/login/student', methods=['GET', 'POST'])
def login_student():
    if request.method == 'POST':
        fname, lname = request.form.get('first_name').strip(), request.form.get('last_name').strip()
        year_group = int(request.form.get('year_group'))
        target_end_year = get_academic_end_year() if year_group == 13 else get_academic_end_year() + 1
        
        student = Student.query.join(Cohort).filter(Student.first_name.ilike(fname), Student.last_name.ilike(lname), Cohort.end_year == target_end_year).first()
        if student:
            login_user(student)
            return redirect(url_for('student_view', id=student.id))
        flash('Student not found in that Year Group. Please check spelling.', 'error')
    return render_template('login_student.html')

@app.route('/login/teacher', methods=['GET', 'POST'])
def login_teacher():
    if request.method == 'POST':
        email, password = request.form.get('email').strip().lower(), request.form.get('password')
        user = Teacher.query.filter_by(email=email).first()
        if user:
            if user.password is None:
                user.password = generate_password_hash(password)
                db.session.commit()
                login_user(user)
                flash('First-time login successful. Your password has been set.', 'success')
                return redirect(url_for('dashboard'))
            elif check_password_hash(user.password, password):
                login_user(user)
                return redirect(url_for('dashboard'))
        flash('Invalid credentials.', 'error')
    return render_template('login_teacher.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('home'))

# --- MAIN DASHBOARDS ---
@app.route('/dashboard')
@teacher_required
def dashboard():
    cohorts = Cohort.query.order_by(Cohort.end_year.desc()).all()
    cohort_data = [{'id': c.id, 'title': get_cohort_title(c.end_year), 'years': f"{c.start_year}-{c.end_year}", 'student_count': len(c.students)} for c in cohorts]
    return render_template('dashboard.html', cohorts=cohort_data)

@app.route('/teachers')
@teacher_required
def teacher_management():
    teachers = Teacher.query.all()
    return render_template('teacher_management.html', teachers=teachers)

@app.route('/curriculum')
@teacher_required
def curriculum_management():
    skills = Skill.query.order_by(Skill.name).all()
    practicals = Practical.query.all()
    return render_template('curriculum_management.html', skills=skills, practicals=practicals)

# --- COHORT ROUTES ---
@app.route('/cohort/add', methods=['POST'])
@teacher_required
def add_cohort():
    db.session.add(Cohort(start_year=int(request.form.get('start_year')), end_year=int(request.form.get('end_year'))))
    db.session.commit()
    flash('Cohort added successfully!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/cohort/delete/<int:id>', methods=['POST'])
@teacher_required
def delete_cohort(id):
    cohort = Cohort.query.get_or_404(id)
    db.session.delete(cohort)
    db.session.commit()
    flash('Cohort and all its students deleted successfully.', 'success')
    return redirect(url_for('dashboard'))

# --- DYNAMIC SKILLS & PRACTICALS ROUTES ---
@app.route('/skill/add', methods=['POST'])
@teacher_required
def add_skill():
    name = request.form.get('name').strip()
    desc = request.form.get('description').strip()
    if Skill.query.filter_by(name=name).first():
        flash('Skill already exists.', 'error')
    else:
        db.session.add(Skill(name=name, description=desc))
        db.session.commit()
        flash('Skill added.', 'success')
    return redirect(url_for('curriculum_management'))

@app.route('/skill/edit/<int:id>', methods=['POST'])
@teacher_required
def edit_skill(id):
    skill = Skill.query.get_or_404(id)
    skill.name = request.form.get('name').strip()
    skill.description = request.form.get('description').strip()
    db.session.commit()
    flash('Skill updated.', 'success')
    return redirect(url_for('curriculum_management'))

@app.route('/skill/delete/<int:id>', methods=['POST'])
@teacher_required
def delete_skill(id):
    skill = Skill.query.get_or_404(id)
    Assessment.query.filter_by(skill_id=id).delete()
    db.session.delete(skill)
    db.session.commit()
    flash('Skill deleted.', 'success')
    return redirect(url_for('curriculum_management'))

@app.route('/practical/add', methods=['POST'])
@teacher_required
def add_practical():
    name = request.form.get('name').strip()
    description = request.form.get('description', '').strip()
    skill_ids = request.form.getlist('skills')
    
    p = Practical(name=name, description=description)
    if skill_ids:
        selected_skills = Skill.query.filter(Skill.id.in_(skill_ids)).all()
        p.skills.extend(selected_skills)
        
    db.session.add(p)
    db.session.commit()
    flash('Practical added.', 'success')
    return redirect(url_for('curriculum_management'))

@app.route('/practical/edit/<int:id>', methods=['POST'])
@teacher_required
def edit_practical(id):
    p = Practical.query.get_or_404(id)
    p.name = request.form.get('name').strip()
    p.description = request.form.get('description', '').strip()
    
    skill_ids = request.form.getlist('skills')
    p.skills = []
    if skill_ids:
        selected_skills = Skill.query.filter(Skill.id.in_(skill_ids)).all()
        p.skills.extend(selected_skills)
        
    db.session.commit()
    flash('Practical updated.', 'success')
    return redirect(url_for('curriculum_management'))

@app.route('/practical/delete/<int:id>', methods=['POST'])
@teacher_required
def delete_practical(id):
    p = Practical.query.get_or_404(id)
    Assessment.query.filter_by(practical_id=id).delete()
    Attendance.query.filter_by(practical_id=id).delete()
    db.session.delete(p)
    db.session.commit()
    flash('Practical deleted.', 'success')
    return redirect(url_for('curriculum_management'))

@app.route('/teacher/add', methods=['POST'])
@teacher_required
def add_teacher():
    email = request.form.get('email').strip().lower()
    if Teacher.query.filter_by(email=email).first():
        flash('Email already registered to a teacher.', 'error')
    else:
        db.session.add(Teacher(
            title=request.form.get('title'),
            first_name=request.form.get('first_name').strip(),
            last_name=request.form.get('last_name').strip(),
            email=email,
            password=None 
        ))
        db.session.commit()
        flash('Teacher added. They can set their password on their first login.', 'success')
    return redirect(url_for('teacher_management'))

@app.route('/teacher/edit/<int:id>', methods=['POST'])
@teacher_required
def edit_teacher(id):
    teacher = Teacher.query.get_or_404(id)
    teacher.title = request.form.get('title')
    teacher.first_name = request.form.get('first_name').strip()
    teacher.last_name = request.form.get('last_name').strip()
    teacher.email = request.form.get('email').strip().lower()
    
    new_pass = request.form.get('new_password')
    if new_pass:
        teacher.password = generate_password_hash(new_pass)
    db.session.commit()
    flash(f"Updated details for {teacher.title} {teacher.last_name}.", 'success')
    return redirect(url_for('teacher_management'))

@app.route('/teacher/delete/<int:id>', methods=['POST'])
@teacher_required
def delete_teacher(id):
    if current_user.id == id:
        flash("You cannot delete yourself.", "error")
        return redirect(url_for('teacher_management'))
    teacher = Teacher.query.get_or_404(id)
    db.session.delete(teacher)
    db.session.commit()
    flash("Teacher deleted.", "success")
    return redirect(url_for('teacher_management'))

# --- COHORT & STUDENT VIEWS ---
@app.route('/cohort/<int:id>')
@teacher_required
def cohort_view(id):
    cohort = Cohort.query.get_or_404(id)
    is_y13 = get_cohort_status(cohort)
    practicals, all_skills = Practical.query.all(), Skill.query.all()
    matrix = [{'id': p.id, 'name': p.name, 'skills': [s.name for s in p.skills]} for p in practicals]

    students_data = []
    unique_classes = set()
    
    for s in cohort.students:
        current_class = s.y13_class if is_y13 else s.y12_class
        if current_class:
            unique_classes.add(current_class)
            
        # Determine Display string (e.g. "13D (was 12B)")
        display_class = current_class or "Unassigned"
        if is_y13 and s.y12_class and s.y13_class != s.y12_class:
            display_class = f"{s.y13_class} (was {s.y12_class})"
        elif is_y13 and not s.y13_class and s.y12_class:
            display_class = f"Unassigned (was {s.y12_class})"

        assessments = Assessment.query.filter_by(student_id=s.id).all()
        skill_totals = {sk.name: 0 for sk in all_skills}
        prac_totals = {p.id: 0 for p in practicals}
        assessed_cells = {p.id: {} for p in practicals}
        
        for a in assessments:
            sk_name = Skill.query.get(a.skill_id).name
            skill_totals[sk_name] += 1
            prac_totals[a.practical_id] += 1
            assessed_cells[a.practical_id][sk_name] = True
            
        status = "Fail" if any(count < 3 for count in skill_totals.values()) else "Pass"
        students_data.append({
            'id': s.id, 'first_name': s.first_name, 'last_name': s.last_name, 
            'current_class': current_class or "", 'display_class': display_class,
            'total_ticks': len(assessments), 'status': status, 
            'assessed_cells': assessed_cells
        })

    return render_template('cohort.html', cohort=cohort, is_y13=is_y13, unique_classes=sorted(list(unique_classes)), 
                           title=get_cohort_title(cohort.end_year), matrix=json.dumps(matrix), 
                           students_json=json.dumps(students_data))

@app.route('/cohort/<int:id>/add_student', methods=['POST'])
@teacher_required
def add_student(id):
    cohort = Cohort.query.get_or_404(id)
    is_y13 = get_cohort_status(cohort)
    cname = request.form.get('class_name', '').strip()
    
    y12_c = cname if not is_y13 else None
    y13_c = cname if is_y13 else None
    
    db.session.add(Student(
        first_name=request.form.get('first_name').strip(), 
        last_name=request.form.get('last_name').strip(), 
        cohort_id=id, y12_class=y12_c, y13_class=y13_c
    ))
    db.session.commit()
    flash('Student added.', 'success')
    return redirect(url_for('cohort_view', id=id))

@app.route('/cohort/<int:id>/bulk_add', methods=['POST'])
@teacher_required
def bulk_add(id):
    cohort = Cohort.query.get_or_404(id)
    is_y13 = get_cohort_status(cohort)
    format_type = request.form.get('format')
    data = request.form.get('students_data').strip().split('\n')
    
    added_count = 0
    for line in data:
        line = line.strip()
        if not line: continue
        fname, lname, cname = parse_student_line(line, format_type)
        if fname and lname:
            y12_c = cname if not is_y13 else None
            y13_c = cname if is_y13 else None
            db.session.add(Student(first_name=fname, last_name=lname, cohort_id=id, y12_class=y12_c, y13_class=y13_c))
            added_count += 1
            
    db.session.commit()
    flash(f'Successfully imported {added_count} students.', 'success')
    return redirect(url_for('cohort_view', id=id))

@app.route('/cohort/<int:id>/bulk_update_classes', methods=['POST'])
@teacher_required
def bulk_update_classes(id):
    cohort = Cohort.query.get_or_404(id)
    is_y13 = get_cohort_status(cohort)
    format_type = request.form.get('format')
    data = request.form.get('students_data').strip().split('\n')
    
    updated_count = 0
    for line in data:
        line = line.strip()
        if not line: continue
        fname, lname, cname = parse_student_line(line, format_type)
        if fname and lname and cname:
            # Locate student by first and last name within this cohort
            student = Student.query.filter(Student.cohort_id == id, Student.first_name.ilike(fname), Student.last_name.ilike(lname)).first()
            if student:
                if is_y13:
                    student.y13_class = cname
                else:
                    student.y12_class = cname
                updated_count += 1
                
    db.session.commit()
    flash(f'Successfully updated classes for {updated_count} existing students.', 'success')
    return redirect(url_for('cohort_view', id=id))

@app.route('/cohort/<int:id>/delete_students', methods=['POST'])
@teacher_required
def delete_students(id):
    student_ids = request.form.getlist('student_ids')
    if student_ids:
        Student.query.filter(Student.id.in_(student_ids)).delete(synchronize_session=False)
        db.session.commit()
        flash(f'Deleted {len(student_ids)} student(s).', 'success')
    return redirect(url_for('cohort_view', id=id))

# --- BULK MARK & API ROUTES ---
@app.route('/cohort/<int:cohort_id>/practical/<int:prac_id>/bulk_mark')
@teacher_required
def bulk_mark(cohort_id, prac_id):
    cohort = Cohort.query.get_or_404(cohort_id)
    practical = Practical.query.get_or_404(prac_id)
    is_y13 = get_cohort_status(cohort)
    
    # Pre-fetch all data to pass to the template cleanly
    attendances = Attendance.query.filter(Attendance.student_id.in_([s.id for s in cohort.students]), Attendance.practical_id == practical.id).all()
    att_dict = {a.student_id: a.is_present for a in attendances}
    
    assessments = Assessment.query.filter(Assessment.student_id.in_([s.id for s in cohort.students]), Assessment.practical_id == practical.id).all()
    assessed_dict = {s.id: [] for s in cohort.students}
    for a in assessments:
        assessed_dict[a.student_id].append(a.skill_id)
        
    unique_classes = set()
    for s in cohort.students:
        c = s.y13_class if is_y13 else s.y12_class
        if c: unique_classes.add(c)
            
    sorted_students = sorted(cohort.students, key=lambda s: s.last_name)
    return render_template('bulk_mark.html', cohort=cohort, practical=practical, students=sorted_students, 
                           skills=practical.skills, att_dict=att_dict, assessed_dict=assessed_dict, 
                           is_y13=is_y13, unique_classes=sorted(list(unique_classes)))

@app.route('/api/update_mark', methods=['POST'])
@teacher_required
def api_update_mark():
    data = request.get_json()
    student_ids = data.get('student_ids', [])
    prac_id = data.get('prac_id')
    mark_type = data.get('type') 
    skill_id = data.get('skill_id')
    is_achieved = data.get('value')

    for sid in student_ids:
        if mark_type == 'attendance':
            att = Attendance.query.filter_by(student_id=sid, practical_id=prac_id).first()
            if is_achieved and not att:
                db.session.add(Attendance(student_id=sid, practical_id=prac_id, teacher_id=current_user.id, is_present=True))
            elif att:
                att.is_present = is_achieved
                att.teacher_id = current_user.id
                att.timestamp = datetime.utcnow()
        elif mark_type == 'skill':
            ass = Assessment.query.filter_by(student_id=sid, practical_id=prac_id, skill_id=skill_id).first()
            if is_achieved and not ass:
                db.session.add(Assessment(student_id=sid, practical_id=prac_id, skill_id=skill_id, teacher_id=current_user.id))
            elif not is_achieved and ass:
                db.session.delete(ass)
                
    db.session.commit()
    return jsonify({'success': True})

# --- INDIVIDUAL STUDENT VIEWS ---
@app.route('/student/<int:id>')
@login_required
def student_view(id):
    if current_user.role == 'student' and current_user.id != id: abort(403)
    student = Student.query.get_or_404(id)
    is_y13 = get_cohort_status(student.cohort)
    
    assessments = Assessment.query.filter_by(student_id=student.id).all()
    attendances = Attendance.query.filter_by(student_id=student.id).all()
    
    prac_scores = {p: 0 for p in Practical.query.all()}
    skill_scores = {sk: 0 for sk in Skill.query.all()}
    attendance_status = {a.practical_id: a.is_present for a in attendances}
    
    for a in assessments:
        prac_scores[Practical.query.get(a.practical_id)] += 1
        skill_scores[Skill.query.get(a.skill_id)] += 1
        
    return render_template('student.html', student=student, prac_scores=prac_scores, skill_scores=skill_scores, 
                           total_score=len(assessments), date=datetime.utcnow().strftime('%d/%m/%Y'), 
                           attendance_status=attendance_status, is_y13=is_y13)

@app.route('/grade/<int:student_id>/<int:prac_id>', methods=['GET', 'POST'])
@login_required
def grade(student_id, prac_id):
    if current_user.role == 'student' and current_user.id != student_id: abort(403)
    student = Student.query.get_or_404(student_id)
    practical = Practical.query.get_or_404(prac_id)
    is_y13 = get_cohort_status(student.cohort)
    skills = practical.skills
    
    if request.method == 'POST' and current_user.role == 'teacher':
        is_present = request.form.get('attendance_present') == 'on'
        att_record = Attendance.query.filter_by(student_id=student.id, practical_id=practical.id).first()
        
        if is_present and not att_record:
            db.session.add(Attendance(student_id=student.id, practical_id=practical.id, teacher_id=current_user.id, is_present=True))
        elif is_present and att_record and not att_record.is_present:
            att_record.is_present = True
            att_record.teacher_id = current_user.id
            att_record.timestamp = datetime.utcnow()
        elif not is_present and att_record and att_record.is_present:
            att_record.is_present = False
            att_record.teacher_id = current_user.id
            att_record.timestamp = datetime.utcnow()

        for sk in skills:
            achieved = request.form.get(f'skill_{sk.id}') == 'on'
            existing = Assessment.query.filter_by(student_id=student.id, skill_id=sk.id, practical_id=practical.id).first()
            if achieved and not existing:
                db.session.add(Assessment(student_id=student.id, skill_id=sk.id, practical_id=practical.id, teacher_id=current_user.id))
            elif not achieved and existing:
                db.session.delete(existing)
                
        db.session.commit()
        flash('Updated Successfully', 'success')
        return redirect(url_for('cohort_view', id=student.cohort_id))

    assessments = Assessment.query.filter_by(student_id=student.id, practical_id=practical.id).all()
    assessed_skill_ids = {a.skill_id: a for a in assessments}
    
    signatures = {}
    for a in assessments:
        t = Teacher.query.get(a.teacher_id)
        signatures[a.skill_id] = f"{t.title} {t.first_name[0]} {t.last_name} ({a.date_signed.strftime('%d/%m/%Y')})"

    att_record = Attendance.query.filter_by(student_id=student.id, practical_id=practical.id).first()
    att_signature = ""
    if att_record and att_record.is_present:
        t = Teacher.query.get(att_record.teacher_id)
        att_signature = f"{t.title} {t.first_name[0]} {t.last_name} ({att_record.timestamp.strftime('%d/%m/%Y')})"

    return render_template('grade.html', student=student, practical=practical, skills=skills, assessed=assessed_skill_ids, 
                           signatures=signatures, attendance=att_record, att_signature=att_signature, is_y13=is_y13)

@app.route('/export_data')
@teacher_required
def export_data():
    wb = Workbook()
    
    ws_teachers = wb.active
    ws_teachers.title = "Teachers"
    ws_teachers.append(["ID", "Title", "First Name", "Last Name", "Email", "Account Status"])
    for t in Teacher.query.all():
        status = "Active" if t.password else "Pending Setup"
        ws_teachers.append([t.id, t.title, t.first_name, t.last_name, t.email, status])
        
    ws_cohorts = wb.create_sheet("Cohorts")
    ws_cohorts.append(["ID", "Start Year", "End Year"])
    for c in Cohort.query.all():
        ws_cohorts.append([c.id, c.start_year, c.end_year])

    ws_students = wb.create_sheet("Students")
    ws_students.append(["ID", "First Name", "Last Name", "Cohort ID", "Y12 Class", "Y13 Class"])
    for s in Student.query.all():
        ws_students.append([s.id, s.first_name, s.last_name, s.cohort_id, s.y12_class or "", s.y13_class or ""])

    ws_skills = wb.create_sheet("Skills")
    ws_skills.append(["ID", "Skill Name", "Description"])
    for sk in Skill.query.all():
        ws_skills.append([sk.id, sk.name, sk.description])

    ws_practicals = wb.create_sheet("Practicals")
    ws_practicals.append(["ID", "Practical Name", "Description"])
    for p in Practical.query.all():
        ws_practicals.append([p.id, p.name, p.description])

    ws_assessments = wb.create_sheet("Assessments_Log")
    ws_assessments.append(["Record ID", "Student ID", "Student Name", "Skill ID", "Practical ID", "Teacher ID", "Date Signed"])
    for a in Assessment.query.all():
        student = Student.query.get(a.student_id)
        student_name = f"{student.last_name}, {student.first_name}" if student else "Unknown"
        date_str = a.date_signed.strftime('%d/%m/%Y %H:%M') if a.date_signed else ""
        ws_assessments.append([a.id, a.student_id, student_name, a.skill_id, a.practical_id, a.teacher_id, date_str])

    ws_attendance = wb.create_sheet("Attendance_Log")
    ws_attendance.append(["Record ID", "Student ID", "Practical ID", "Teacher ID", "Is Present", "Timestamp"])
    for a in Attendance.query.all():
        ws_attendance.append([a.id, a.student_id, a.practical_id, a.teacher_id, a.is_present, a.timestamp.strftime('%d/%m/%Y %H:%M') if a.timestamp else ""])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"Tiffin_CPAC_Database_Export_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(
        out, 
        as_attachment=True, 
        download_name=filename, 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --- SEEDING & STARTUP ---
def seed_database():
    if not Teacher.query.first():
        teachers = [
            ("Mrs", "Ann", "Noble", "ANoble@tiffin.kingston.sch.uk"),
            ("Mr", "Kurt", "Braganza", "KBraganza@tiffin.kingston.sch.uk"),
            ("Dr", "Matteo", "Bocchi", "MBocchi@tiffin.kingston.sch.uk"),
            ("Dr", "Payal", "Tyagi", "PTyagi@tiffin.kingston.sch.uk"),
            ("Mr", "Tom", "Wightwick", "TWightwick@tiffin.kingston.sch.uk"),
            ("Mr", "Dhaval", "Pandey", "6060@tiffin.kingston.sch.uk")
        ]
        for t, f, l, e in teachers:
            db.session.add(Teacher(title=t, first_name=f, last_name=l, email=e.lower(), password=None))
        db.session.commit()

    if not Skill.query.first():
        skill_definitions = {
            '1a': "Correctly follows written instructions...", '2a': "Correctly uses appropriate instrumentation...", '2b': "Carries out techniques methodically...",
            '2c': "Identifies and controls significant quantitative variables...", '2d': "Selects appropriate equipment...", '3a': "Identifies hazards and assesses risks...",
            '3b': "Uses appropriate safety equipment...", '4a': "Makes accurate observations...", '4b': "Obtains accurate data and records methodically...",
            '5a': "Uses appropriate software to process data...", '5b': "Cites sources of information..."
        }
        skill_objs = {k: Skill(name=k, description=v) for k, v in skill_definitions.items()}
        for s in skill_objs.values(): db.session.add(s)
        db.session.commit()

        practical_mappings = [
            ("1. Stationary Waves", ['1a', '2c', '3a', '3b', '4a', '4b']), ("2.a Young's slit", ['2a', '2b', '3a', '3b', '4a', '4b']),
            ("2.b Diffraction Gratings", ['2a', '2b', '2c', '2d', '4a', '4b']), ("3. Determination of g", ['1a', '2a', '2d', '4b', '5a', '5b']),
            ("4. Young modulus", ['1a', '2c', '2d', '3b', '4a', '5a', '5b']), ("5. Resistivity of a wire", ['2a', '2c', '3b', '4a', '5a', '5b']),
            ("6. emf", ['2a', '2c', '3b', '4a', '5a', '5b']), ("7.a SHM - Simple Pendulum", ['1a', '2a', '2b', '4b', '5a', '5b']),
            ("7.b SHM - Mass spring", ['1a', '2a', '2b', '4b']), ("8.a Boyle's Law", ['1a', '2b', '2c', '2d', '3a', '4b']),
            ("8b Charles' Law", ['1a', '2c', '3a', '3b', '4a', '5a', '5b']), ("9.a Capacitor Discharging", ['1a', '2c', '2d', '3a', '3b', '4a']),
            ("9.b Capacitor Charging", ['1a', '3b', '4a', '4b']), ("10. Motor Effect", ['2b', '2d', '3a', '3b', '4a', '4b']),
            ("11. Search Coils", ['1a', '2a', '2c', '2d', '4b']), ("12. Gamma Radiation", ['1a', '2b', '2d', '3a', '3b', '4a', '4b'])
        ]
        for p_name, s_keys in practical_mappings:
            p = Practical(name=p_name)
            for sk in s_keys: p.skills.append(skill_objs[sk])
            db.session.add(p)
        db.session.commit()

with app.app_context():
    db.create_all()
    seed_database()

if __name__ == '__main__':
    app.run(debug=True, port=5000)